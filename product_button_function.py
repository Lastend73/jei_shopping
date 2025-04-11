import csv
import db_mysql as get_data
from PySide6.QtWidgets import (QApplication, QHeaderView, QMainWindow, QPushButton,
    QSizePolicy, QStatusBar, QTabWidget, QTableWidget, 
    QTableWidgetItem, QWidget, QFileDialog, QMessageBox)
import time
import openpyxl
import os
from openpyxl.styles import Alignment
import traceback
import re

product_path = "info\품목정보.csv"

def on_pushButton_clicked(table_widget, result_button=None): #품목 조회 버튼
    with open(product_path, mode='r', encoding='cp949') as file:

        db_product_list = get_data.get_product_data()

        table_widget.setRowCount(0)

        reader = csv.reader(file)
        header = next(reader)
        product_name =[]
        product_info = []
        
        for datas in reader:
            row = table_widget.rowCount()
            product_name.append(datas[1])
            if datas[2] == "" or datas[3] == "":
                table_widget.insertRow(row)
                for column,data in enumerate(datas) :
                    item = QTableWidgetItem(str(data))
                    table_widget.setItem(row, column, item)


        for result in db_product_list:
            if result[1] not in product_name :
                row = table_widget.rowCount()
                table_widget.insertRow(row)
                for column,data in enumerate(result) :
                    item = QTableWidgetItem(str(data))
                    table_widget.setItem(row, column, item)

        # 매칭이 필요한 값이 없는지 확인
        needs_mapping = False
        for row in range(table_widget.rowCount()):
            if table_widget.item(row, 2) is None or table_widget.item(row, 2).text() == "":
                needs_mapping = True
                break

        # 매칭이 필요한 값이 없고 결과 버튼이 전달된 경우 결과 버튼 활성화
        if not needs_mapping and result_button is not None:
            result_button.setEnabled(True)

def on_pushButton_5_clicked(): #데이터 출력
    product_info_list =[]
    package_list = []
    unmatched_set = set()
    unmatched_rows = []
    matched_rows = []

    with open(product_path, mode='r', encoding='cp949') as file:
        csv_datas= csv.reader(file)
        next(csv_datas)
        for csv_data in csv_datas :
            product_info_list.append(csv_data)

    folder_path = 'result'
    if folder_path:
        results = get_data.get_order_data()
        point_use_list = get_data.get_point_use_list()

        file_path = os.path.join("info\판매오더일괄등록_템플릿.xlsx") # 템플릿 파일 경로
        try :
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                header = ["웹주문번호", "고객코드", "오더일", "납품예정일", "영업사원명", "판매구분", "출하유형" , "납품처주소", "결제유형", "카드사정보", "품목코드", "수량", "단가", "통화", "공급가액", "부가세", "총액", "판매유형", "수령방법", "보증개월수", "무상여부", "비고", "가격정책예외"]
                sheet.append(header)

            delete_number_list_check = []
            for result in results:
                if int(result[12]) == 0 :
                    delete_number_list_check.append(result[0])
                for product_info in product_info_list:
                    if result[10] == product_info[1] and (product_info[3] == "x" or product_info[3] == "X"):
                        delete_number_list_check.append(result[0])
                        package_list.append(result[0])
                        continue
            
            for point_use in point_use_list: # 포인트 사용 내역 삭제
                delete_number_list_check.append(point_use[0])
            
            delete_number_list = []
            for result in results:
                if int(result[0]) in delete_number_list_check:
                    delete_number_list.append(int(result[0]))
                    continue
            # 모든 result 항목을 리스트로 변환
            for i in range(len(results)):
                results[i] = list(results[i])
            print(results)
            row_num = 2  # 데이터가 시작될 행 번호 (헤더 다음 행)
            transaction_number_dic = {} #거래번호 리스트

            # 삭제할 항목을 먼저 처리
            results_to_keep = []
            for result in results:
                if result[0] in delete_number_list:
                    transaction_number_dic[result[0]] = result[23]
                else:
                    results_to_keep.append(result)
            
            results = results_to_keep  # 삭제되지 않은 항목들만 남김

            for result in results:
                # 비고 필드(result[21])에서 휴대폰 번호와 비고 분리
                if '/' in result[21]:
                    phone_num, note = result[21].split('/', 1)  # '/'를 기준으로 첫 번째 발생 지점에서 분리
                    # 숫자만 추출하고 하이픈(-) 추가
                    phone_num = re.sub(r'\D', '', phone_num)  # 숫자 외의 모든 문자 제거
                    if len(phone_num) == 11:  # 휴대폰 번호 형식(11자리)
                        phone_num = phone_num[:3] + '-' + phone_num[3:7] + '-' + phone_num[7:]
                    elif len(phone_num) == 10:  # 일반 전화번호 형식(10자리)
                        phone_num = phone_num[:3] + '-' + phone_num[3:6] + '-' + phone_num[6:]
                else:
                    phone_num = ""
                    note = result[21]
                if result[8] == "신용카드":
                    result[8] = "T1"
                    result[21] = f"쇼핑몰 신용카드 주문 건\n거래명세서 포함 택배요청\n배송요청: {note}\n연락처: {phone_num}\n배송지 :{result[7]}"
                if result[8] == "무통장입금":
                    result[8] = "T2"
                    result[21] = f"쇼핑몰 무통장 주문 건\n거래명세서 포함 택배요청\n배송요청: {note}\n연락처: {phone_num}\n배송지 :{result[7]}"

                if not re.search(r'[a-zA-Z가-힣]',  result[1]): 
                    result[1]=re.sub(r'\D', '', result[1])
                    if len(result[1]) == 10 :
                        result[1] = result[1][:3] + '-' + result[1][3:5] + '-' + result[1][5:]
                    else:
                        result[1] = result[1] + " (미등록)"
                else:
                    result[1] = result[1] + " (미등록)"

                product_check_index = -1
                for product_info in product_info_list:
                    if result[10] == product_info[1]:
                        result[11] = int(result[11])  # 수량
                        result[14] = int(result[14])  # 공급가액
                        result[16] = int(result[16]) 
                        if product_info[2] != "":
                            result[10] = product_info[2]
                        if product_info[3] != "" :
                            product_check_index = 1
                            if int(product_info[3]) > 1 :
                                result[12] = int(result[12]) / int(product_info[3])  # 단가
                                result[11] = int(result[11]) * int(product_info[3])  # 수량
                                result[14] = int(result[12] / 11 * 10) * result[11]  # 공급가액
                                result[15] = int(result[12] / 11) * result[11]  # 부가세
                                result[16] = int(result[12]) * result[11]  # 총액
                                break
                if product_check_index == -1:
                    result[10] = result[10] + " (미등록)"
                    unmatched_set.add(result[0])

                if result[0] in unmatched_set:
                    unmatched_rows.append(result)
                else:
                    matched_rows.append(result)    

            results = unmatched_rows + matched_rows
            for r in results:
                if r[0] == 25174:
                    print(r)
            for result in results:
                for col_num, value in enumerate(result[:-1], 1):  # 마지막 열을 제외하고 작성
                    sheet.cell(row=row_num, column=col_num, value=value)
                row_num += 1

            for row in sheet.iter_rows(min_row=1):
                cell = row[21]
                cell.alignment = Alignment(wrap_text=True)  # 비고란 자동 줄바꿈
                if "(미등록)" in row[10].value :
                    row[10].font = openpyxl.styles.Font(color="FF0000")
                    row[10].alignment = Alignment(vertical='center')  
                # 고객코드(2번째 열)에 하이픈이 없고 미등록 표시가 있는 경우 빨간색으로 표시
                if row[1].value and not '-' in str(row[1].value) and "(미등록)" in str(row[1].value):
                    row[1].font = openpyxl.styles.Font(color="FF0000")
                    row[1].alignment = Alignment(vertical='center')
                for i in range(2,4) : 
                    row[i].alignment = Alignment(horizontal='center',vertical='center')  # 오더일 납품예정일 중앙정렬
                for i in [12, 14, 15, 16]:
                    row[i].number_format = '#,##0'
                    row[i].alignment = Alignment(horizontal='right',vertical='center')  # 수량, 공급가액, 부가세, 총액 오른쪽 정렬
            
            workbook.save(f'{folder_path}/{time.strftime("%Y_%m_%d_%H_%M")}_판매오더.xlsx')
            memo = open(f'{folder_path}/{time.strftime("%Y_%m_%d_%H_%M")}_패키지_주문리스트.txt', 'w')

            for key, value in transaction_number_dic.items():
                memo.write(f"{key} : {value}\n")
            memo.close()

            msgBox = QMessageBox()
            msgBox.setWindowTitle("다운로드 완료")  # 메시지창의 상단 제목
            msgBox.setIcon(QMessageBox.Information)  # 메시지창 내부에 표시될 아이콘
            msgBox.setText("다운로드 완료")  # 메시지 제지
            msgBox.exec()
        except :
            msgBox = QMessageBox()
            msgBox.setWindowTitle("Alert Window") # 메세지창의 상단 제목
            msgBox.setIcon(QMessageBox.Information) # 메세지창 내부에 표시될 아이콘
            msgBox.setText("문제발생") # 메세지 제목
            msgBox.exec()
            print("An error occurred:")
            print(traceback.format_exc())
    else:
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Alert Window") # 메세지창의 상단 제목
        msgBox.setIcon(QMessageBox.Information) # 메세지창 내부에 표시될 아이콘
        msgBox.setText("다운로드 경로가 선택되지 않았습니다.") # 메세지 제목

def on_pushButton_6_clicked(table_widget, result_button): #품목 수정
    row_count = table_widget.rowCount()
    column_count = table_widget.columnCount()

    fixed_row_datas=[]
    for row in range(row_count):
        row_data_tem = []
        for column in range(column_count):
            item = table_widget.item(row, column) 
            if item is not None:
                row_data_tem.append(item.text())
            else:
                row_data_tem.append('')
        fixed_row_datas.append(row_data_tem)
    
    csv_row_datas = []
    with open(product_path, mode='r', encoding='cp949') as file:
        csv_datas= csv.reader(file)
        for csv_data in csv_datas :
            csv_row_datas.append(csv_data)

    for j in range(len(fixed_row_datas)):
        match_found = False
        for i in range(len(csv_row_datas)):
            if csv_row_datas[i][0] == fixed_row_datas[j][0] and csv_row_datas[i][1] == fixed_row_datas[j][1]:
                csv_row_datas[i] = fixed_row_datas[j]
                match_found = True
                break
        if not match_found :
            csv_row_datas.append(fixed_row_datas[j])

    with open(product_path, 'w', newline='', encoding='cp949') as file:
        writer = csv.writer(file)
        for row in csv_row_datas:
            writer.writerow(row)

    on_pushButton_clicked(table_widget)
    
    # 모든 품목이 맵핑되었는지 확인
    all_mapped = True
    for row in range(table_widget.rowCount()):
        if table_widget.item(row, 2) is None or table_widget.item(row, 2).text() == "":
            all_mapped = False
            break
    
    # 모든 품목이 맵핑되었다면 결과 출력 버튼 활성화
    if all_mapped:
        result_button.setEnabled(True)
    else:
        result_button.setEnabled(False)

def get_table_data(table_widget):
    row_count = table_widget.rowCount()
    column_count = table_widget.columnCount()
    table_data = []

    for row in range(row_count):
        row_data = []
        for column in range(column_count):
            item = table_widget.item(row, column)
            if item is not None:
                row_data.append(item.text())
            else:
                row_data.append('')
        table_data.append(row_data)

    return table_data
